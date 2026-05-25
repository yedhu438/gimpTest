"""
export_orders_with_images.py
Query local DB for up to 10 orders that have at least one customer image,
and export full details to Excel.

Run:
    python export_orders_with_images.py
Output:
    C:\\Varsany\\Output\\orders_with_images_YYYYMMDD_HHMMSS.xlsx
"""

import os
import sys
from datetime import datetime

import pyodbc
import pandas as pd

from db import get_connection as _db_get_connection

# ─── CONFIG ───────────────────────────────────────────────────────────────────

OUTPUT_DIR  = r"C:\Varsany\Output"
OUTPUT_FILE = os.path.join(
    OUTPUT_DIR,
    f"orders_with_images_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

LIMIT = 10

# ─── DATABASE ─────────────────────────────────────────────────────────────────

def _connect():
    return _db_get_connection()


def fetch_image_orders(limit=LIMIT):
    conn = _connect()
    cur  = conn.cursor()

    sql = f"""
        SELECT TOP {limit}
            o.OrderID,
            o.SKU,
            o.ItemType,
            o.Quantity,
            o.IsShipped,
            CONVERT(nvarchar(36), d.idCustomOrderDetails) AS idCustomOrderDetails,
            d.PrintLocation,
            d.IsDesignComplete,
            d.IsOrderProcess,
            d.FrontImage,  d.FrontPreviewImage,
            d.FrontText,   d.FrontFonts,   d.FrontColours,
            d.BackImage,   d.BackPreviewImage,
            d.BackText,    d.BackFonts,    d.BackColours,
            d.PocketImage, d.PocketPreviewImage,
            d.PocketText,  d.PocketFonts,  d.PocketColours,
            d.SleeveImage, d.SleevePreviewImage,
            d.SleeveText
        FROM tblCustomOrder o
        JOIN tblCustomOrderDetails d ON o.idCustomOrder = d.idCustomOrder
        WHERE (
            ISNULL(d.FrontImage,  '') <> '' OR
            ISNULL(d.BackImage,   '') <> '' OR
            ISNULL(d.PocketImage, '') <> '' OR
            ISNULL(d.SleeveImage, '') <> ''
        )
        ORDER BY o.DateAdd ASC
    """
    cur.execute(sql)
    cols = [c[0] for c in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    conn.close()
    return rows


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def active_image_zones(row):
    zones = []
    for zone in ("Front", "Back", "Pocket", "Sleeve"):
        val = row.get(f"{zone}Image") or ""
        if str(val).strip():
            zones.append(zone)
    return ", ".join(zones)


# ─── EXPORT ───────────────────────────────────────────────────────────────────

def export(rows):
    if not rows:
        print("No orders with images found in the database.")
        return None

    df = pd.DataFrame(rows)

    # Insert summary column at the front
    df.insert(0, "ImageZones", df.apply(active_image_zones, axis=1))

    # Clean up boolean-looking columns
    for col in ("IsShipped", "IsDesignComplete", "IsOrderProcess"):
        if col in df.columns:
            df[col] = df[col].apply(lambda v: "Yes" if v else "No")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Orders With Images")

        ws = writer.sheets["Orders With Images"]

        # Auto-fit column widths (capped at 80 for URL columns)
        for col_cells in ws.columns:
            width = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(width + 2, 80)

        # Bold header row
        from openpyxl.styles import Font as XLFont
        for cell in ws[1]:
            cell.font = XLFont(bold=True)

        # Make image URL cells clickable hyperlinks
        from openpyxl.styles import Font as XLFont
        from openpyxl.styles import colors
        link_font = XLFont(color="0000FF", underline="single")
        image_cols = {
            cell.value: cell.column
            for cell in ws[1]
            if cell.value and "Image" in str(cell.value)
        }
        for row in ws.iter_rows(min_row=2):
            for col_name, col_idx in image_cols.items():
                cell = row[col_idx - 1]
                url  = str(cell.value or "").strip()
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.font = link_font

    print(f"\nExported {len(df)} order(s) with images: {OUTPUT_FILE}")
    return OUTPUT_FILE


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("Connecting to DB...")
    try:
        rows = fetch_image_orders(LIMIT)
    except pyodbc.Error as e:
        print(f"DB connection failed: {e}")
        sys.exit(1)

    print(f"Found {len(rows)} order(s) with images.\n")
    for i, r in enumerate(rows, 1):
        zones = active_image_zones(r)
        print(f"  {i}. OrderID={r['OrderID']}  SKU={r['SKU']}  "
              f"ImageZones=[{zones}]  PrintLocation={r['PrintLocation']}")

    export(rows)


if __name__ == "__main__":
    main()
