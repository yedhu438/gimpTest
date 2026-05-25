"""
export_recent_orders.py
Export all orders added in the last 2 hours from the live database to Excel.
"""

import os
import sys
from datetime import datetime

import pyodbc
import pandas as pd

from db import get_connection as _db_get_connection

OUTPUT_DIR  = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(
    OUTPUT_DIR,
    f"orders_recent_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)


def fetch_recent_orders():
    conn = _db_get_connection()
    cur  = conn.cursor()
    sql = """
        SELECT
            o.OrderID,
            o.SKU,
            o.ItemType,
            o.Quantity,
            o.Gender,
            o.BuyerName,
            o.PurchaseDate,
            o.ShipByDate,
            o.IsShipped,
            o.Notes,
            o.DateAdd,
            CONVERT(nvarchar(36), d.idCustomOrderDetails) AS idCustomOrderDetails,
            d.PrintLocation,
            d.IsFrontLocation,
            d.IsBackLocation,
            d.IsPocketLocation,
            d.IsSleeveLocation,
            d.FrontText,   d.FrontFonts,   d.FrontColours,
            d.FrontImage,  d.FrontPreviewImage,
            d.BackText,    d.BackFonts,    d.BackColours,
            d.BackImage,   d.BackPreviewImage,
            d.PocketText,  d.PocketFonts,  d.PocketColours,
            d.PocketImage, d.PocketPreviewImage,
            d.SleeveText,  d.SleeveFonts,  d.SleeveColours,
            d.SleeveImage, d.SleevePreviewImage,
            d.IsOrderProcess,
            d.IsDesignComplete,
            d.ProcessBy,
            d.ProcessTime
        FROM tblCustomOrder o
        JOIN tblCustomOrderDetails d ON o.idCustomOrder = d.idCustomOrder
        WHERE o.DateAdd >= DATEADD(HOUR, -2, GETUTCDATE())
        ORDER BY o.DateAdd DESC
    """
    cur.execute(sql)
    cols = [c[0] for c in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    conn.close()
    return rows


def export(rows):
    if not rows:
        print("No orders found in the last 2 hours.")
        return None

    df = pd.DataFrame(rows)

    bool_cols = ["IsShipped", "IsOrderProcess", "IsDesignComplete",
                 "IsFrontLocation", "IsBackLocation", "IsPocketLocation", "IsSleeveLocation"]
    for col in bool_cols:
        if col in df.columns:
            df[col] = df[col].apply(lambda v: "Yes" if v else "No")

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Recent Orders")
        ws = writer.sheets["Recent Orders"]

        from openpyxl.styles import Font as XLFont, PatternFill, Alignment

        # Bold + coloured header
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = XLFont(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Clickable hyperlinks on image/preview URL columns
        link_font = XLFont(color="0000FF", underline="single")
        image_cols = {
            cell.value: cell.column
            for cell in ws[1]
            if cell.value and ("Image" in str(cell.value) or "Preview" in str(cell.value))
        }
        for row in ws.iter_rows(min_row=2):
            for col_name, col_idx in image_cols.items():
                cell = row[col_idx - 1]
                url = str(cell.value or "").strip()
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.font = link_font

        # Auto-fit column widths
        for col_cells in ws.columns:
            width = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(width + 2, 60)

        # Freeze header row
        ws.freeze_panes = "A2"

    print(f"\nExported {len(df)} order(s) to: {OUTPUT_FILE}")
    return OUTPUT_FILE


def main():
    print("Connecting to live database...")
    try:
        rows = fetch_recent_orders()
    except pyodbc.Error as e:
        print(f"DB connection failed: {e}")
        sys.exit(1)

    print(f"Found {len(rows)} order(s) from the last 2 hours.\n")
    for i, r in enumerate(rows, 1):
        print(f"  {i}. OrderID={r['OrderID']}  SKU={r['SKU']}  "
              f"PrintLocation={r['PrintLocation']}  DesignComplete={r['IsDesignComplete']}")

    export(rows)


if __name__ == "__main__":
    main()
