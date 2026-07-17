"""
Import both sheets of the test1 Excel into local DB, then kick off the batch processor.
"""
import openpyxl, pyodbc, os, subprocess, sys
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

from db import get_connection as _db_get_connection

EXCEL_PATH = r"C:\gimpTest\Output\test1\UnshippedDTFOrders_04052026_014331.xlsx"
OUTPUT_DIR = r"C:\gimpTest\Output\test1 Export"
DATE_AFTER = "2026-04-30"

def b(v):
    if v is True:  return 1
    if v is False: return 0
    return None

def run():
    print("Reading Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    o_ws = wb["DTF Orders"]
    d_ws = wb["DTF Order Details"]
    o_hdrs = [c.value for c in next(o_ws.iter_rows(min_row=1, max_row=1))]
    d_hdrs = [c.value for c in next(d_ws.iter_rows(min_row=1, max_row=1))]
    o_rows = list(o_ws.iter_rows(min_row=2, values_only=True))
    d_rows = list(d_ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    print(f"  DTF Orders rows       : {len(o_rows)}")
    print(f"  DTF Order Details rows: {len(d_rows)}")

    conn = _db_get_connection()
    cur  = conn.cursor()

    # ── tblCustomOrder ────────────────────────────────────────────────────────
    ins_o = skp_o = 0
    for row in o_rows:
        d = dict(zip(o_hdrs, row))
        cur.execute("SELECT COUNT(*) FROM tblCustomOrder WHERE idCustomOrder=?", d["idCustomOrder"])
        if cur.fetchone()[0]:
            skp_o += 1
            continue
        cols = [
            "idCustomOrder","OrderID","OrderItemID","ASIN","SKU","Quantity","ItemImageUrl",
            "Gender","ItemType","PurchaseDate","ConvertedPurchaseDate","BuyerName","PostalCode",
            "SalesChannel","Notes","IsCustomOrderDetailsGet","CustomOrderDetailsGetTime",
            "AlertEmailProcessTime","DateAdd","IsShipped","ShippedStatusSetTime","IsNotesUpdated",
            "NotesUpdateTime","ShipByDate","ConvertedShipByDate","UpdatedDate","IsItemTypeGet",
            "ItemUpdatedTime","RepeatOrderId",
        ]
        vals = [
            d["idCustomOrder"],
            d["OrderID"],
            str(d["OrderItemID"]) if d["OrderItemID"] is not None else None,
            d["ASIN"], d["SKU"], d["Quantity"], d["ItemImageUrl"],
            d["Gender"], d["ItemType"], d["PurchaseDate"], d["ConvertedPurchaseDate"],
            d["BuyerName"], d["PostalCode"], d["SalesChannel"], d["Notes"],
            b(d["IsCustomOrderDetailsGet"]), d["CustomOrderDetailsGetTime"],
            d["AlertEmailProcessTime"], d["DateAdd"],
            b(d["IsShipped"]), d["ShippedStatusSetTime"],
            b(d["IsNotesUpdated"]), d["NotesUpdateTime"],
            d["ShipByDate"], d["ConvertedShipByDate"], d["UpdatedDate"],
            b(d["IsItemTypeGet"]), d["ItemUpdatedTime"], d["RepeatOrderId"],
        ]
        ph = ",".join(["?"] * len(vals))
        cur.execute(f"INSERT INTO tblCustomOrder ({','.join(cols)}) VALUES ({ph})", vals)
        ins_o += 1
    conn.commit()
    print(f"tblCustomOrder       : {ins_o} inserted, {skp_o} already existed")

    # ── tblCustomOrderDetails ─────────────────────────────────────────────────
    ins_d = skp_d = 0
    for row in d_rows:
        d = dict(zip(d_hdrs, row))
        cur.execute("SELECT COUNT(*) FROM tblCustomOrderDetails WHERE idCustomOrderDetails=?", d["idCustomOrderDetails"])
        if cur.fetchone()[0]:
            cur.execute("UPDATE tblCustomOrderDetails SET IsDesignComplete=0 WHERE idCustomOrderDetails=?",
                        d["idCustomOrderDetails"])
            skp_d += 1
            continue
        cols = [
            "idCustomOrderDetails","idCustomOrder","Title","PrintLocation",
            "FrontLabel","FrontPreviewImage","FrontImage","FrontImageJSON",
            "FrontFonts","FrontColours","FrontText","FrontTextJSON","FrontPSD","IsFrontPSDDownload",
            "BackLabel","BackPreviewImage","BackImage","BackImageJSON",
            "BackFonts","BackColours","BackText","BackTextJSON","BackPSD","IsBackPSDDownload",
            "PocketLabel","PocketPreviewImage","PocketImage","PocketImageJSON",
            "PocketFonts","PocketColours","PocketText","PocketTextJSON","PocketPSD","IsPocketPSDDownload",
            "SleeveLabel","SleevePreviewImage","SleeveImage","SleeveImageJSON",
            "SleeveFonts","SleeveColours","SleeveText","SleeveTextJSON","SleevePSD","IsSleevePSDDownload",
            "CustomizationJSON","CustomizationJSONTemp","AdditionalPSD","IsAdditionalPSDDownload",
            "ConfirmOrderLabel","ConfirmOrderValue","ConfirmOrderPreviewImage",
            "IsFrontLocation","IsBackLocation","IsPocketLocation","IsSleeveLocation",
            "IsOrderClick","ProcessBy","ProcessTime","IsOrderProcess","DateAdd",
            "IsDesignComplete","IsOrderItemIdUpdated",
        ]
        vals = [
            d["idCustomOrderDetails"], d["idCustomOrder"], d.get("Title"), d.get("PrintLocation"),
            d.get("FrontLabel"), d.get("FrontPreviewImage"), d.get("FrontImage"), d.get("FrontImageJSON"),
            d.get("FrontFonts"), d.get("FrontColours"), d.get("FrontText"), d.get("FrontTextJSON"),
            d.get("FrontPSD"), b(d.get("IsFrontPSDDownload")),
            d.get("BackLabel"), d.get("BackPreviewImage"), d.get("BackImage"),
            None,  # BackImageJSON not in Excel export
            d.get("BackFonts"), d.get("BackColours"), d.get("BackText"), d.get("BackTextJSON"),
            d.get("BackPSD"), b(d.get("IsBackPSDDownload")),
            d.get("PocketLabel"), d.get("PocketPreviewImage"), d.get("PocketImage"), d.get("PocketImageJSON"),
            d.get("PocketFonts"), d.get("PocketColours"), d.get("PocketText"), d.get("PocketTextJSON"),
            d.get("PocketPSD"), b(d.get("IsPocketPSDDownload")),
            d.get("SleeveLabel"), d.get("SleevePreviewImage"), d.get("SleeveImage"), d.get("SleeveImageJSON"),
            d.get("SleeveFonts"), d.get("SleeveColours"), d.get("SleeveText"), d.get("SleeveTextJSON"),
            d.get("SleevePSD"), b(d.get("IsSleevePSDDownload")),
            d.get("CustomizationJSON"), d.get("CustomizationJSONTemp"),
            d.get("AdditionalPSD"), b(d.get("IsAdditionalPSDDownload")),
            d.get("ConfirmOrderLabel"), d.get("ConfirmOrderValue"), d.get("ConfirmOrderPreviewImage"),
            b(d.get("IsFrontLocation")), b(d.get("IsBackLocation")),
            b(d.get("IsPocketLocation")), b(d.get("IsSleeveLocation")),
            b(d.get("IsOrderClick")), d.get("ProcessBy"), d.get("ProcessTime"),
            b(d.get("IsOrderProcess")), d.get("DateAdd"),
            0,  # IsDesignComplete = 0 → batch processor will pick this up
            b(d.get("IsOrderItemIdUpdated")),
        ]
        assert len(cols) == len(vals), f"Column/value count mismatch: {len(cols)} cols vs {len(vals)} vals"
        ph = ",".join(["?"] * len(vals))
        cur.execute(f"INSERT INTO tblCustomOrderDetails ({','.join(cols)}) VALUES ({ph})", vals)
        ins_d += 1
    conn.commit()
    conn.close()
    print(f"tblCustomOrderDetails: {ins_d} inserted, {skp_d} reset to IsDesignComplete=0")

    print(f"\nRunning batch processor (date-after {DATE_AFTER}) → {OUTPUT_DIR}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    result = subprocess.run(
        [sys.executable, "batch_processor.py",
         "--date-after", DATE_AFTER,
         "--output",     OUTPUT_DIR],
        cwd=os.path.dirname(__file__)
    )
    print(f"\nBatch processor exited with code {result.returncode}")

if __name__ == "__main__":
    run()
